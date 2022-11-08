import openpyxl

def clean_file():
    path = "waslabels.xlsx"
    results = open("artikel_aantal_waslabels.txt", "w")

    wb_obj = openpyxl.load_workbook(path)

    # Set which sheet. Original file can only have one sheet
    sheet_obj = wb_obj.active

    # Checks the amount of rows and columns. Original file can only have 4 columns.
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    # Keeps track of how many lines are written.
    count = 0

    # Iterate through every row and concatenates a string which will be written to a csv-file.
    for x in range(1, m_row + 1):
        try:
            str_val = ""

            for y in range(1, m_col + 1):
                cell_obj = sheet_obj.cell(row = x, column=y)
                if cell_obj.value == None:
                    str_val += ""
                else:
                    if y < m_col:
                        str_val += str(cell_obj.value) + "\t"
                    else:
                        str_val += str(cell_obj.value)

            # Checks if it is the last row. Empty rows raise an error, so the last row can't be followed by an enter.
            if x < m_row:
                results.writelines(str_val + "\n")
            else:
                results.writelines(str_val)
            count += 1
        except:
            print(f"Error parsing row {x}, column {y}")

    # count-1 because of the headers. These don't count towards the final amount.
    print(f"Done writing {count-1} lines.")


if __name__ == "__main__":
    clean_file()
